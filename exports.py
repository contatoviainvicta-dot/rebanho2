"""
exports.py — Exportação de relatórios PDF e Excel.

Funções:
  gerar_excel_lote(lote_id, animais, pesagens_dict, ocorrencias_dict) -> bytes
  gerar_pdf_relatorio(titulo, secoes) -> bytes
    onde secoes = [{"titulo": str, "df": pd.DataFrame}, ...]
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    _REPORTLAB = True
except ImportError:
    _REPORTLAB = False


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

def _estilo_cabecalho(ws, row, cols, fill_hex="1F5C2E"):
    fill = PatternFill("solid", fgColor=fill_hex)
    bold = Font(bold=True, color="FFFFFF", size=11)
    alin = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = bold
        cell.alignment = alin


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def gerar_excel_lote(nome_lote: str, animais: list, pesagens_dict: dict,
                      ocorrencias_dict: dict) -> bytes:
    """
    Gera um arquivo Excel com abas: Resumo, Animais, Pesagens, Ocorrências.

    Parâmetros:
      animais          → lista de tuplas (id, identificacao, idade, lote_id)
      pesagens_dict    → {animal_id: [(id, animal_id, peso, data), ...]}
      ocorrencias_dict → {animal_id: [(id, animal_id, data, tipo, desc,
                                       grav, custo, dias, status), ...]}
    """
    wb = Workbook()

    # ---- ABA: RESUMO ----
    ws_r = wb.active
    ws_r.title = "Resumo"
    ws_r["A1"] = f"Relatório — {nome_lote}"
    ws_r["A1"].font = Font(bold=True, size=14)
    ws_r["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_r["A2"].font = Font(italic=True, size=10)

    total_animais = len(animais)
    total_ocorrencias = sum(len(v) for v in ocorrencias_dict.values())
    custo_sanitario = sum(
        o[6] for ocs in ocorrencias_dict.values() for o in ocs if o[6]
    )

    # GMD médio
    gmds = []
    for aid, pesos in pesagens_dict.items():
        if len(pesos) > 1:
            df = pd.DataFrame(pesos, columns=["id", "aid", "peso", "data"])
            df["data"] = pd.to_datetime(df["data"])
            df = df.sort_values("data")
            dias = (df["data"].iloc[-1] - df["data"].iloc[0]).days
            if dias > 0:
                gmd = (df["peso"].iloc[-1] - df["peso"].iloc[0]) / dias
                if 0 <= gmd <= 2:
                    gmds.append(gmd)
    gmd_medio = sum(gmds) / len(gmds) if gmds else 0

    resumo = [
        ["Indicador", "Valor"],
        ["Total de animais", total_animais],
        ["Total de ocorrências", total_ocorrencias],
        ["Custo sanitário total (R$)", f"{custo_sanitario:.2f}"],
        ["GMD médio (kg/dia)", f"{gmd_medio:.3f}"],
    ]
    for i, linha in enumerate(resumo, start=4):
        for j, val in enumerate(linha, start=1):
            ws_r.cell(row=i, column=j, value=val)
    _estilo_cabecalho(ws_r, 4, 2)
    _auto_width(ws_r)

    # ---- ABA: ANIMAIS ----
    ws_a = wb.create_sheet("Animais")
    cab_a = ["ID", "Identificação", "Idade (meses)", "Lote ID"]
    for j, c in enumerate(cab_a, 1):
        ws_a.cell(row=1, column=j, value=c)
    _estilo_cabecalho(ws_a, 1, len(cab_a))
    for i, animal in enumerate(animais, start=2):
        for j, val in enumerate(animal, start=1):
            ws_a.cell(row=i, column=j, value=val)
    _auto_width(ws_a)

    # ---- ABA: PESAGENS ----
    ws_p = wb.create_sheet("Pesagens")
    cab_p = ["ID", "Animal ID", "Peso (kg)", "Data"]
    for j, c in enumerate(cab_p, 1):
        ws_p.cell(row=1, column=j, value=c)
    _estilo_cabecalho(ws_p, 1, len(cab_p))
    row = 2
    for pesos in pesagens_dict.values():
        for p in pesos:
            for j, val in enumerate(p, start=1):
                ws_p.cell(row=row, column=j, value=val)
            row += 1
    _auto_width(ws_p)

    # ---- ABA: OCORRÊNCIAS ----
    ws_o = wb.create_sheet("Ocorrências")
    cab_o = ["ID", "Animal ID", "Data", "Tipo", "Descrição",
             "Gravidade", "Custo (R$)", "Dias Recuperação", "Status"]
    for j, c in enumerate(cab_o, 1):
        ws_o.cell(row=1, column=j, value=c)
    _estilo_cabecalho(ws_o, 1, len(cab_o))
    row = 2
    for ocs in ocorrencias_dict.values():
        for o in ocs:
            for j, val in enumerate(o, start=1):
                ws_o.cell(row=row, column=j, value=val)
            row += 1
    _auto_width(ws_o)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_excel_sanitario(vacinas: list, medicamentos: list) -> bytes:
    """Excel com abas de agenda sanitária e estoque de medicamentos."""
    wb = Workbook()

    ws_v = wb.active
    ws_v.title = "Agenda Vacinas"
    cab_v = ["ID", "Lote ID", "Vacina", "Previsto", "Realizado", "Status", "Observação"]
    for j, c in enumerate(cab_v, 1):
        ws_v.cell(row=1, column=j, value=c)
    _estilo_cabecalho(ws_v, 1, len(cab_v), "0F6E56")
    for i, v in enumerate(vacinas, start=2):
        for j, val in enumerate(v, start=1):
            ws_v.cell(row=i, column=j, value=val)
    _auto_width(ws_v)

    ws_m = wb.create_sheet("Estoque Medicamentos")
    cab_m = ["ID", "Nome", "Unidade", "Estoque Atual", "Estoque Mínimo", "Validade", "Custo Unit. (R$)"]
    for j, c in enumerate(cab_m, 1):
        ws_m.cell(row=1, column=j, value=c)
    _estilo_cabecalho(ws_m, 1, len(cab_m), "0F6E56")
    for i, m in enumerate(medicamentos, start=2):
        for j, val in enumerate(m, start=1):
            ws_m.cell(row=i, column=j, value=val)
    _auto_width(ws_m)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def gerar_pdf_relatorio(titulo: str, secoes: list) -> bytes:
    """
    Gera PDF com título e uma ou mais seções, cada uma com um DataFrame.

    secoes = [
        {"titulo": "Animais do lote", "df": df_animais},
        {"titulo": "Ocorrências",      "df": df_oc},
    ]
    Retorna bytes do PDF.
    Se reportlab não estiver disponível, retorna PDF mínimo de fallback.
    """
    if not _REPORTLAB:
        return _pdf_fallback(titulo, secoes)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    # Título
    style_titulo = ParagraphStyle(
        "titulo", parent=styles["Title"],
        fontSize=16, spaceAfter=6, textColor=colors.HexColor("#1F5C2E"),
    )
    story.append(Paragraph(titulo, style_titulo))
    story.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#1F5C2E")))
    story.append(Spacer(1, 0.4 * cm))

    style_sec = ParagraphStyle(
        "sec", parent=styles["Heading2"],
        fontSize=12, textColor=colors.HexColor("#1F5C2E"), spaceAfter=4,
    )

    for sec in secoes:
        story.append(Paragraph(sec["titulo"], style_sec))
        df: pd.DataFrame = sec["df"]

        if df is None or df.empty:
            story.append(Paragraph("Sem dados.", styles["Normal"]))
            story.append(Spacer(1, 0.3 * cm))
            continue

        # cabeçalho + dados
        data = [list(df.columns)] + df.astype(str).values.tolist()
        col_w = (doc.width) / len(df.columns)
        t = Table(data, colWidths=[col_w] * len(df.columns), repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1F5C2E")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 9),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F1F8F3")]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5 * cm))

    doc.build(story)
    return buf.getvalue()


def _pdf_fallback(titulo: str, secoes: list) -> bytes:
    """PDF mínimo em texto puro quando reportlab não está disponível."""
    linhas = [f"% {titulo}", f"% {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""]
    for sec in secoes:
        linhas.append(f"## {sec['titulo']}")
        df = sec.get("df")
        if df is not None and not df.empty:
            linhas.append(df.to_string(index=False))
        linhas.append("")
    return "\n".join(linhas).encode()
